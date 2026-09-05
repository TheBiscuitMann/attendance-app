# Roster import and eligibility export.
#
# Kept out of views.py because these endpoints deal in files, not JSON:
# the import view reads an uploaded spreadsheet, the export view writes
# one. Everything else about them (auth, teacher scoping) matches the
# rest of the API.

import base64
import csv
import io
import re
from datetime import date

from django.db import transaction
from rest_framework import status
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Batch, Student
from .views import ELIGIBILITY_THRESHOLD, build_summary

# A roster sheet is tiny. Anything bigger than this is the wrong file.
MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB
MAX_PDF_PAGES = 40
MAX_ROWS = 2000

NAVY = '0B2A59'
RED = 'D32F2F'


# ── Reading the uploaded sheet ────────────────────────────────────────────────

def _clean_cell(value):
    """Turn whatever the sheet holds into a trimmed string.

    Excel loves storing '231115081' as the float 231115081.0 — undo
    that, or every imported ID grows a spurious '.0'.
    """
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _rows_from_xlsx(uploaded):
    from openpyxl import load_workbook
    workbook = load_workbook(uploaded, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = []
    for row in sheet.iter_rows(max_row=MAX_ROWS, values_only=True):
        rows.append([_clean_cell(cell) for cell in row])
    workbook.close()
    return rows


def _rows_from_csv(uploaded):
    # utf-8-sig eats the BOM Excel prepends when it saves CSV.
    text = uploaded.read().decode('utf-8-sig', errors='replace')
    reader = csv.reader(io.StringIO(text))
    return [[_clean_cell(cell) for cell in row] for row in reader][:MAX_ROWS]


def _normalize_header(cell):
    return re.sub(r'[^a-z]', '', cell.lower())


ID_HEADERS = {
    'id', 'studentid', 'sid', 'reg', 'regno', 'regid', 'registration',
    'registrationno', 'registrationid', 'roll', 'rollno', 'rollnumber',
}


def _looks_like_id(value):
    # Registration numbers are digits with optional separators:
    # 231-115-081, 2021/33/104, 20103115081.
    return bool(value) and bool(re.fullmatch(r'[\d\s\-\/\._]+', value))


def _detect_columns(rows):
    """Work out which column is the ID and whether row 1 is a header.

    Returns (id_col, name_col, first_data_row_index).
    """
    first = rows[0]
    id_col, name_col = None, None
    for index, cell in enumerate(first[:6]):
        normalized = _normalize_header(cell)
        if normalized in ID_HEADERS and id_col is None:
            id_col = index
        elif 'name' in normalized and name_col is None:
            name_col = index

    if id_col is not None or name_col is not None:
        # A header row exists. Fill in whichever side it didn't name.
        if id_col is None:
            id_col = 1 if name_col == 0 else 0
        if name_col is None:
            name_col = 1 if id_col == 0 else 0
        return id_col, name_col, 1

    # No header. The agreed format is ID first, name second — but check
    # a sample and swap if the sheet clearly came the other way round.
    sample = [row for row in rows[:20] if len(row) >= 2 and (row[0] or row[1])]
    if sample:
        first_ids = sum(_looks_like_id(row[0]) for row in sample)
        second_ids = sum(_looks_like_id(row[1]) for row in sample)
        if second_ids > first_ids:
            return 1, 0, 0
    return 0, 1, 0


# ── Reading a PDF roster ──────────────────────────────────────────────────────
#
# PDFs are a page-layout format, not a data format: there are no cells,
# only text at coordinates. So extraction is best-effort, which is why
# the PDF path always previews before writing anything.

def _is_date_like(token):
    """True for 05/09/2026 and 2026-09-05, false for 231-115-081.

    Dates in headers and footers otherwise get mistaken for IDs.
    A four-digit year on either end is the giveaway.
    """
    match = re.fullmatch(r'(\d{1,4})[\-/](\d{1,2})[\-/](\d{2,4})', token)
    if not match:
        return False
    first, _, last = match.groups()
    return len(first) == 4 or len(last) == 4


SERIAL_PREFIX = re.compile(r'^\s*\d{1,3}\s*[\.\)]\s+')
ID_TOKEN = re.compile(r'(?<![\w\-])(\d[\d\-\/\._]{2,}\d)(?![\w])')


def _looks_like_student_id(cell, min_digits=4):
    """In a table the column gives context, so 4 digits is enough.
    In free text there is no column, so callers ask for more — a bare
    "2026" in a header line is otherwise indistinguishable from an ID.
    """
    return (
        bool(cell)
        and bool(re.fullmatch(r'[\d\s\-\/\._]+', cell))
        and sum(ch.isdigit() for ch in cell) >= min_digits
        and not _is_date_like(cell)
    )


def _score_pdf_columns(rows):
    """Pick the ID and name columns of a PDF table by their contents.

    A PDF roster usually has a serial column ("1, 2, 3…") that a
    position-based guess would mistake for the ID. Scoring by content
    instead: the ID column is the one with long digit strings, the name
    column the one with words.
    """
    width = max(len(row) for row in rows)
    id_scores = [0] * width
    name_scores = [0] * width

    for row in rows:
        for index in range(width):
            cell = row[index] if index < len(row) else ''
            if not cell:
                continue
            if _looks_like_student_id(cell):
                id_scores[index] += 1
            elif re.search(r'[A-Za-z]', cell) and not any(ch.isdigit() for ch in cell):
                name_scores[index] += 1

    if not any(id_scores) or not any(name_scores):
        return None, None

    id_col = id_scores.index(max(id_scores))
    ranked = sorted(range(width), key=lambda i: name_scores[i], reverse=True)
    name_col = next((i for i in ranked if i != id_col and name_scores[i] > 0), None)
    return (id_col, name_col) if name_col is not None else (None, None)


def _has_header(row):
    for cell in row[:6]:
        normalized = _normalize_header(cell)
        if normalized in ID_HEADERS or 'name' in normalized:
            return True
    return False


def _pairs_from_pdf_tables(page_tables):
    """Extract (student_id, name) from tables pdfplumber found."""
    rows = []
    for table in page_tables:
        for row in table:
            cleaned = [_clean_cell(cell) for cell in row]
            if any(cleaned):
                rows.append(cleaned)

    if not rows or max(len(row) for row in rows) < 2:
        return []

    id_col, name_col = _score_pdf_columns(rows)
    if id_col is None:
        return []

    start = 1 if _has_header(rows[0]) else 0
    pairs = []
    for number, row in enumerate(rows[start:], start=start + 1):
        student_id = row[id_col] if id_col < len(row) else ''
        name = row[name_col] if name_col < len(row) else ''
        if _looks_like_student_id(student_id) and re.search(r'[A-Za-z]', name):
            pairs.append((number, student_id, name))
    return pairs


def _pairs_from_pdf_text(text, start_number=1):
    """Fallback: pull an ID and a name out of each text line.

    Handles "1. 231-115-081  Ahsanul Haque" and the same line without a
    serial or with the columns reversed.
    """
    pairs = []
    number = start_number
    for raw_line in text.split('\n'):
        line = SERIAL_PREFIX.sub('', raw_line.strip())
        if not line:
            continue

        found = None
        for match in ID_TOKEN.finditer(line):
            if _looks_like_student_id(match.group(1), min_digits=5):
                found = match
                break
        if not found:
            continue

        name = (line[:found.start()] + ' ' + line[found.end():])
        name = re.sub(r'\s{2,}', ' ', name).strip(' .-|\t')
        # Reject anything that reads as a label or a sentence rather than
        # a name: headers ("Session: Fall 2026"), footers, prose. A colon
        # anywhere is the strongest signal; so is an implausible length.
        if (not name
                or ':' in name
                or len(name.split()) > 6
                or not re.search(r'[A-Za-z]', name)):
            continue

        pairs.append((number, found.group(1), name))
        number += 1
    return pairs


def _pairs_from_pdf(uploaded):
    """Rows from a PDF: tables first, text lines as the fallback.

    Raises ValueError with a teacher-readable message when the file
    can't yield a roster.
    """
    import pdfplumber

    table_pairs = []
    text_chunks = []
    with pdfplumber.open(uploaded) as pdf:
        if len(pdf.pages) > MAX_PDF_PAGES:
            raise ValueError(
                f'That PDF has {len(pdf.pages)} pages. Please upload a roster '
                f'of at most {MAX_PDF_PAGES} pages.'
            )
        for page in pdf.pages:
            try:
                tables = page.extract_tables() or []
            except Exception:
                tables = []
            if tables:
                table_pairs.extend(_pairs_from_pdf_tables(tables))
            text_chunks.append(page.extract_text() or '')

    if table_pairs:
        return table_pairs, 'pdf-table'

    full_text = '\n'.join(text_chunks).strip()
    if not full_text:
        raise ValueError(
            'This PDF has no readable text — it looks like a scan or a photo. '
            'Scanned rosters cannot be imported reliably; please upload the '
            'list as .xlsx or .csv instead.'
        )

    text_pairs = _pairs_from_pdf_text(full_text)
    if not text_pairs:
        raise ValueError(
            'No student rows could be found in that PDF. It needs a student ID '
            'and a name on each line. If the layout is unusual, save the list '
            'as .xlsx or .csv instead.'
        )
    return text_pairs, 'pdf-text'


# ── Turning parsed rows into students ─────────────────────────────────────────

def _clean_rows(rows):
    """Drop rows where every cell is empty."""
    return [row for row in rows if any(cell for cell in row)]


def _pairs_from_sheet(rows):
    """(row_number, student_id, name) from a spreadsheet's rows."""
    id_col, name_col, start = _detect_columns(rows)
    pairs = []
    for number, row in enumerate(rows[start:], start=start + 1):
        student_id = row[id_col] if id_col < len(row) else ''
        name = row[name_col] if name_col < len(row) else ''
        pairs.append((number, student_id, name))
    return pairs


def _classify(batch, pairs):
    """Sort parsed rows into new / already-enrolled / duplicate / invalid."""
    existing_ids = set(batch.students.values_list('student_id', flat=True))
    seen_in_file = set()

    new_rows = []
    skipped_existing = []
    duplicate_rows = []
    invalid_rows = []

    for number, student_id, name in pairs:
        student_id = (student_id or '').strip()[:50]
        name = (name or '').strip()[:100]
        entry = {'row': number, 'student_id': student_id, 'name': name}

        if not student_id or not name:
            invalid_rows.append(entry)
        elif student_id in existing_ids:
            skipped_existing.append(entry)
        elif student_id in seen_in_file:
            duplicate_rows.append(entry)
        else:
            seen_in_file.add(student_id)
            new_rows.append(entry)

    return {
        'new_rows': new_rows,
        'skipped_existing': skipped_existing,
        'duplicates_in_file': duplicate_rows,
        'invalid_rows': invalid_rows,
        'total_rows': len(pairs),
    }


class ImportStudentsView(APIView):
    """POST /api/students/import/ — bulk-add students to a batch.

    Three shapes, all on this one endpoint:

    * multipart with `file` — parse the sheet or PDF and create the
      students found (the .xlsx / .csv path).
    * multipart with `file` and `preview=1` — parse and return what
      *would* be created, writing nothing. Used for PDFs, where the
      layout is a guess rather than a grid.
    * JSON with `students` — create the rows the teacher confirmed
      after that preview.

    Columns are student_id and name; a header row is optional.
    Students already in the batch are never touched.
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        try:
            batch = Batch.objects.get(id=request.data.get('batch'), teacher=request.user)
        except (Batch.DoesNotExist, ValueError, TypeError):
            return Response({'error': 'Batch not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Confirm step: the rows came back from a preview.
        if 'students' in request.data and not request.FILES.get('file'):
            return self._commit_confirmed(batch, request.data.get('students'))

        return self._handle_upload(batch, request)

    # ── multipart upload ─────────────────────────────────────────────
    def _handle_upload(self, batch, request):
        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'error': 'Attach a file to import.'},
                            status=status.HTTP_400_BAD_REQUEST)
        if uploaded.size > MAX_UPLOAD_BYTES:
            return Response({'error': 'That file is too large. A roster should be '
                                      'well under 5 MB.'},
                            status=status.HTTP_400_BAD_REQUEST)

        name_lower = (uploaded.name or '').lower()
        is_pdf = name_lower.endswith('.pdf')

        try:
            if name_lower.endswith('.xlsx'):
                pairs, source = _pairs_from_sheet(_clean_rows(_rows_from_xlsx(uploaded))), 'xlsx'
            elif name_lower.endswith('.csv'):
                pairs, source = _pairs_from_sheet(_clean_rows(_rows_from_csv(uploaded))), 'csv'
            elif is_pdf:
                pairs, source = _pairs_from_pdf(uploaded)
            else:
                return Response({'error': 'Please upload a .xlsx, .csv or .pdf file. '
                                          'Old .xls files need re-saving as .xlsx.'},
                                status=status.HTTP_400_BAD_REQUEST)
        except ValueError as err:
            # Raised by the PDF reader with a message meant for the teacher.
            return Response({'error': str(err)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception:
            return Response({'error': 'That file could not be read. Re-save it and '
                                      'try again, or use .xlsx instead.'},
                            status=status.HTTP_400_BAD_REQUEST)

        if not pairs:
            return Response({'error': 'The file has no rows to import.'},
                            status=status.HTTP_400_BAD_REQUEST)

        report = _classify(batch, pairs[:MAX_ROWS])

        # PDF layout is inferred, so the teacher checks it before anything
        # is written. Spreadsheets import directly, as before.
        wants_preview = str(request.data.get('preview', '')).lower() in ('1', 'true', 'yes')
        if is_pdf or wants_preview:
            return Response({
                'preview': True,
                'source': source,
                'students': report['new_rows'],
                'skipped_existing': report['skipped_existing'],
                'duplicates_in_file': report['duplicates_in_file'],
                'invalid_rows': report['invalid_rows'],
                'total_rows': report['total_rows'],
            })

        return self._create(batch, report)

    # ── confirmed rows from a preview ────────────────────────────────
    def _commit_confirmed(self, batch, students):
        if not isinstance(students, list) or not students:
            return Response({'error': 'No students were sent to import.'},
                            status=status.HTTP_400_BAD_REQUEST)
        if len(students) > MAX_ROWS:
            return Response({'error': f'At most {MAX_ROWS} students can be imported '
                                      f'at once.'},
                            status=status.HTTP_400_BAD_REQUEST)

        pairs = []
        for index, entry in enumerate(students, start=1):
            if not isinstance(entry, dict):
                return Response({'error': 'Malformed student list.'},
                                status=status.HTTP_400_BAD_REQUEST)
            pairs.append((
                entry.get('row', index),
                str(entry.get('student_id', '')),
                str(entry.get('name', '')),
            ))

        return self._create(batch, _classify(batch, pairs))

    # ── the write itself ─────────────────────────────────────────────
    def _create(self, batch, report):
        to_create = [
            Student(batch=batch, teacher=batch.teacher,
                    student_id=entry['student_id'], name=entry['name'])
            for entry in report['new_rows']
        ]
        with transaction.atomic():
            Student.objects.bulk_create(to_create)

        return Response({
            'created': len(to_create),
            'skipped_existing': report['skipped_existing'],
            'duplicates_in_file': report['duplicates_in_file'],
            'invalid_rows': report['invalid_rows'],
            'total_rows': report['total_rows'],
        })


# ── Exporting the eligibility list ────────────────────────────────────────────

def _sort_key(entry):
    return entry['student_id']


def _export_filename(batch, extension):
    course = batch.course
    raw = f"{course.code}_{batch.name}_eligibility_{date.today().isoformat()}"
    # Whatever survives this is safe in a Content-Disposition header.
    safe = re.sub(r'[^A-Za-z0-9._-]+', '-', raw).strip('-')
    return f"{safe}.{extension}"


def _build_xlsx(batch, summary):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Eligibility'

    navy_fill = PatternFill('solid', fgColor=NAVY)
    red_font = Font(color=RED, bold=True)
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    course = batch.course
    total_sessions = summary[0]['total_sessions'] if summary else 0

    sheet['A1'] = f"{course.code} — {course.name}"
    sheet['A1'].font = Font(bold=True, size=14, color=NAVY)
    batch_label = batch.name + (f" • {batch.section}" if batch.section else '')
    sheet['A2'] = f"Batch: {batch_label}"
    sheet['A3'] = (f"Sessions held: {total_sessions}   •   "
                   f"Eligibility threshold: {ELIGIBILITY_THRESHOLD}%   •   "
                   f"Generated: {date.today().strftime('%d/%m/%Y')}")
    sheet['A3'].font = Font(size=9, color='666666')

    headers = ['Student ID', 'Name', 'Present', 'Late', 'Absent', 'Attendance %', 'Status']
    header_row = 5
    for col, text in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col, value=text)
        cell.fill = navy_fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for row_index, entry in enumerate(sorted(summary, key=_sort_key), start=header_row + 1):
        eligible = entry['rate'] >= ELIGIBILITY_THRESHOLD
        values = [
            entry['student_id'], entry['name'], entry['present'], entry['late'],
            entry['absent'], entry['rate'] / 100,
            'Eligible' if eligible else 'NOT ELIGIBLE',
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col, value=value)
            cell.border = border
            if col >= 3:
                cell.alignment = Alignment(horizontal='center')
            if col == 6:
                cell.number_format = '0%'
            if not eligible and col in (6, 7):
                cell.font = red_font

    widths = [16, 32, 10, 8, 10, 14, 16]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.freeze_panes = f'A{header_row + 1}'

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_pdf(batch, summary):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    navy = colors.HexColor(f'#{NAVY}')
    red = colors.HexColor(f'#{RED}')

    course = batch.course
    total_sessions = summary[0]['total_sessions'] if summary else 0
    batch_label = batch.name + (f" • {batch.section}" if batch.section else '')

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
        title=f"{course.code} eligibility list",
    )

    title_style = ParagraphStyle('title', fontName='Helvetica-Bold',
                                 fontSize=15, textColor=navy, spaceAfter=2)
    meta_style = ParagraphStyle('meta', fontName='Helvetica',
                                fontSize=8.5, textColor=colors.HexColor('#666666'))

    story = [
        Paragraph(f"{course.code} — {course.name}", title_style),
        Paragraph(f"Attendance eligibility list · Batch: {batch_label}", meta_style),
        Paragraph(
            f"Sessions held: {total_sessions} &nbsp;•&nbsp; "
            f"Eligibility threshold: {ELIGIBILITY_THRESHOLD}% &nbsp;•&nbsp; "
            f"Generated: {date.today().strftime('%d/%m/%Y')} &nbsp;•&nbsp; "
            f"Late arrivals count as present.",
            meta_style,
        ),
        Spacer(1, 6 * mm),
    ]

    data = [['Student ID', 'Name', 'Present', 'Late', 'Absent', '%', 'Status']]
    not_eligible_rows = []
    for index, entry in enumerate(sorted(summary, key=_sort_key), start=1):
        eligible = entry['rate'] >= ELIGIBILITY_THRESHOLD
        if not eligible:
            not_eligible_rows.append(index)
        data.append([
            entry['student_id'], entry['name'], str(entry['present']),
            str(entry['late']), str(entry['absent']), f"{entry['rate']}%",
            'Eligible' if eligible else 'NOT ELIGIBLE',
        ])

    table = Table(
        data,
        colWidths=[28 * mm, 62 * mm, 17 * mm, 14 * mm, 16 * mm, 14 * mm, 28 * mm],
        repeatRows=1,
    )
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), navy),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 8.5),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 8.5),
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F6FA')]),
        ('TOPPADDING', (0, 0), (-1, -1), 3.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3.5),
    ]
    for row in not_eligible_rows:
        style.append(('TEXTCOLOR', (5, row), (6, row), red))
        style.append(('FONT', (5, row), (6, row), 'Helvetica-Bold', 8.5))
    table.setStyle(TableStyle(style))

    story.append(table)
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph(
        f"Students below {ELIGIBILITY_THRESHOLD}%: "
        f"{len(not_eligible_rows)} of {len(summary)}. "
        f"Prezence — Metropolitan University.",
        meta_style,
    ))

    document.build(story)
    return buffer.getvalue()


CONTENT_TYPES = {
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'pdf': 'application/pdf',
}


class ExportSummaryView(APIView):
    """GET /api/attendance/export/?batch=<id>&fmt=xlsx|pdf

    The same numbers as /api/attendance/summary/, as a file the
    department can print or forward.

    The file comes back base64-encoded inside a JSON response rather
    than as a raw download. Download managers like IDM watch network
    traffic for file-type responses and hijack them — the fetch gets
    aborted, the app shows a false "cannot reach the server", and the
    file lands outside the app's control. A JSON payload looks like any
    other API response, so nothing intercepts it; the client decodes it
    and saves the file from memory.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        batch_id = request.query_params.get('batch')
        fmt = (request.query_params.get('fmt') or 'xlsx').lower()

        if fmt not in CONTENT_TYPES:
            return Response({'error': 'fmt must be xlsx or pdf.'},
                            status=status.HTTP_400_BAD_REQUEST)
        try:
            batch = Batch.objects.select_related('course').get(
                id=batch_id, teacher=request.user
            )
        except (Batch.DoesNotExist, ValueError, TypeError):
            return Response({'error': 'Batch not found.'}, status=status.HTTP_404_NOT_FOUND)

        summary = build_summary(request.user, batch.id)
        if not summary:
            return Response({'error': 'This batch has no students to export.'},
                            status=status.HTTP_400_BAD_REQUEST)

        content = _build_xlsx(batch, summary) if fmt == 'xlsx' else _build_pdf(batch, summary)

        return Response({
            'filename': _export_filename(batch, fmt),
            'mime': CONTENT_TYPES[fmt],
            'content': base64.b64encode(content).decode('ascii'),
        })